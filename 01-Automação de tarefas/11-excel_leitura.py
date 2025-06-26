from openpyxl import load_workbook # load_workbook carrega o arquivo

# 1 - Lendo workbook e buscando planilha
wb = load_workbook(filename='file/test.xlsx') # ler
planilha = wb['Planilha 1'] #Busca uma planilha

# Acessando um determinado valor
#print(planilha['A2'].value)

# 3 - Interando valores por meio de loop, e por meio desses dados recuperar valores
for i in range(2, 5):
    #print(f"{planilha['A{i}'].value}")
    ano = planilha['A%s' %i].value
    lucro = planilha['B%s' %i].value
    custo = planilha['C%s' %i].value
    print("{0} teve {1} de lucro e {2} de custo".format(ano, lucro, custo))
    #msg = f"Ano {planilha['A{i}'].value} teve {planilha['B{i}'].value} de lucro e {planilha['C{i}'].value} de custo"
    #print(msg)
