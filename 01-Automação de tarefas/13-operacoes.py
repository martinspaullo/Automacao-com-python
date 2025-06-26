from openpyxl import load_workbook
from openpyxl.drawing.image import Image

wb = load_workbook(filename='files/gastos.xlsx')

planilha = wb['Sheet']

valor_total = 0

# 1- Somando os valores
for i in range(2, 10): #Faz a interação e  pega os valores da coluna b
    valor = int(planilha['B%s' %i].value)
    valor_total += valor
    #1 - Criando Fórmula de Soma
planilha['B10'] = valor_total
# wb.save(filename='files/gastos.xlsx')

# 2- Mesclando células
planilha['A11'] = 'Teste'
planilha.merge_cells('A11:B11') #mescla celulas
planilha.unmerge_cells('A11:B11') # desfaz a mesclagem de celulas

# 3- Inserindo Imagens
img = Image('files/bb_preco.png')
planilha.add_image(img, 'A12')

# 4- Deletando Células
planilha.delete_rows(1) #deleta Linhas
planilha.delete_cols(3) #deleta colunas

wb.save(filename='files/gastos_logo.xlsx')
